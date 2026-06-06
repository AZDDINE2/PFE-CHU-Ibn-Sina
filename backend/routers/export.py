"""
routers/export.py — /api/export/*, /api/email/send
"""
import base64
import io
import os
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from core.database import data
from core.data_loader import get_urg, get_soins

router = APIRouter(tags=["export"])


# ── Pydantic model ────────────────────────────────────────────────────────────
class EmailInput(BaseModel):
    to: str
    subject: str = "Rapport CHU Ibn Sina — Dashboard Urgences"
    body: str = "Veuillez trouver ci-joint le rapport automatique du dashboard."
    pdf_base64: Optional[str] = None
    filename: str = "rapport_CHU.pdf"


# ── CSV exports ───────────────────────────────────────────────────────────────
@router.get("/api/export/urgences")
def export_urgences():
    try:
        buf = io.StringIO()
        get_urg().to_csv(buf, index=False, encoding="utf-8-sig")
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=urgences_CHU_Ibn_Sina.csv"},
        )
    except Exception as e:
        raise HTTPException(500, str(e))


@router.get("/api/export/soins")
def export_soins():
    try:
        buf = io.StringIO()
        get_soins().to_csv(buf, index=False, encoding="utf-8-sig")
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=soins_CHU_Ibn_Sina.csv"},
        )
    except Exception as e:
        raise HTTPException(500, str(e))


@router.get("/api/export/etablissements")
def export_etablissements():
    try:
        buf = io.StringIO()
        data["etab"].to_csv(buf, index=False, encoding="utf-8-sig")
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=etablissements_CHU_Ibn_Sina.csv"},
        )
    except Exception as e:
        raise HTTPException(500, str(e))


@router.get("/api/export/excel")
def export_excel():
    try:
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            get_urg().to_excel(writer, sheet_name="Urgences", index=False)
            get_soins().to_excel(writer, sheet_name="Soins", index=False)
            data["etab"].to_excel(writer, sheet_name="Etablissements", index=False)
            wb = writer.book
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for cell in ws[1]:
                    cell.font      = Font(bold=True, color="FFFFFF")
                    cell.fill      = PatternFill("solid", fgColor="1A3BDB")
                    cell.alignment = Alignment(horizontal="center")
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=8)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=CHU_Ibn_Sina_Data.xlsx"},
        )
    except ImportError:
        raise HTTPException(500, detail="openpyxl non installé. Ajoutez-le à requirements.txt.")
    except Exception as e:
        raise HTTPException(500, str(e))


# ── Email ─────────────────────────────────────────────────────────────────────
@router.post("/api/email/send")
def send_email(payload: EmailInput):
    smtp_host = os.environ.get("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.environ.get("SMTP_PORT", 587))
    smtp_user = os.environ.get("SMTP_USER", "")
    smtp_pass = os.environ.get("SMTP_PASSWORD", "")

    if not smtp_user or not smtp_pass:
        raise HTTPException(
            400,
            detail="SMTP non configuré. Définissez SMTP_USER et SMTP_PASSWORD dans les variables d'environnement.",
        )

    try:
        from email.utils import formataddr
        msg            = MIMEMultipart()
        msg["From"]    = formataddr(("CHU Ibn Sina — Urgences", smtp_user))
        msg["To"]      = payload.to
        msg["Subject"] = payload.subject
        msg.attach(MIMEText(payload.body, "plain", "utf-8"))

        if payload.pdf_base64:
            pdf_bytes = base64.b64decode(payload.pdf_base64)
            part      = MIMEBase("application", "octet-stream")
            part.set_payload(pdf_bytes)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{payload.filename}"')
            msg.attach(part)

        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, payload.to, msg.as_string())

        return {"message": f"Email envoyé à {payload.to}"}
    except Exception as e:
        raise HTTPException(500, detail=str(e))
