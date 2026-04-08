"""
Backend FastAPI — CHU Ibn Sina Rabat
PFE BI & Data Science — Azddine 2024/2025
"""

from fastapi import FastAPI, HTTPException, Depends, status, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
import pandas as pd
import numpy as np
import joblib
import os
import io
import hashlib
import smtplib
import base64
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta
from typing import Optional
from sqlalchemy import create_engine, text
from sqlalchemy.pool import StaticPool, NullPool
import time

try:
    import jwt as pyjwt
    JWT_AVAILABLE = True
except ImportError:
    JWT_AVAILABLE = False

ROOT         = os.environ.get('DATA_ROOT', os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
GOLD         = os.path.join(ROOT, 'data', 'gold')
MODELS       = os.path.join(ROOT, 'models')

# SQLite par défaut — chemin absolu dans le volume /app/data
_default_db  = f"sqlite:///{os.path.join(ROOT, 'data', 'chu.db')}"
DATABASE_URL = os.environ.get('DATABASE_URL', _default_db)

# SQLAlchemy engine — SQLite ou PostgreSQL selon DATABASE_URL
engine = None
try:
    if DATABASE_URL.startswith("sqlite"):
        engine = create_engine(
            DATABASE_URL,
            connect_args={"check_same_thread": False, "timeout": 30},
            poolclass=NullPool,
        )
        # WAL mode = lectures simultanées sans blocage
        with engine.connect() as _c:
            _c.execute(text("PRAGMA journal_mode=WAL"))
            _c.execute(text("PRAGMA synchronous=NORMAL"))
    else:
        engine = create_engine(DATABASE_URL, pool_pre_ping=True)
    with engine.connect() as conn:
        conn.execute(text("SELECT 1"))
    db_type = "SQLite" if DATABASE_URL.startswith("sqlite") else "PostgreSQL"
    print(f"{db_type} connecté : {DATABASE_URL}")
except Exception as e:
    print(f"Base de données non disponible : {e} — mode CSV de secours")

app = FastAPI(title="CHU Ibn Sina API", version="1.0.0")

_cors_origins = os.environ.get("CORS_ORIGINS", "*")
_origins = ["*"] if _cors_origins == "*" else _cors_origins.split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=_origins,
    allow_credentials=_cors_origins != "*",
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Données globales ─────────────────────────────────────────
data = {}

def get_urg() -> pd.DataFrame:
    """Lire depuis le cache mémoire (rechargé au startup et après admission)."""
    df = data.get("urg", pd.DataFrame())
    if df.empty and engine:
        reload_urg()
        df = data.get("urg", pd.DataFrame())
    df["Date_Arrivee"] = pd.to_datetime(df["Date_Arrivee"], errors="coerce")
    return df

def get_soins() -> pd.DataFrame:
    """Lire soins depuis le cache mémoire."""
    return data.get("soins", pd.DataFrame())

def reload_urg():
    """Recharge les données depuis SQLite en mémoire."""
    if engine:
        df = pd.read_sql('SELECT * FROM urgences', engine)
        df["Date_Arrivee"] = pd.to_datetime(df["Date_Arrivee"], errors="coerce")
        data["urg"] = df


@app.on_event("startup")
def load_data():
    try:
        # ── PostgreSQL : créer les tables si elles n'existent pas (sans CSV) ─
        if engine:
            with engine.connect() as conn:
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS urgences_bronze (
                        id_urgence    INTEGER PRIMARY KEY AUTOINCREMENT,
                        id_patient    INTEGER,
                        nom_complet   TEXT,
                        age           INTEGER,
                        sexe          TEXT,
                        cin           TEXT DEFAULT '',
                        groupe_sanguin TEXT,
                        antecedents   TEXT,
                        etablissement TEXT,
                        type_etab     TEXT,
                        ville         TEXT,
                        date_arrivee  TEXT,
                        date_sortie   TEXT,
                        niveau_triage TEXT,
                        motif_consultation TEXT,
                        orientation   TEXT,
                        duree_sejour_min INTEGER,
                        nb_medecins_dispo INTEGER,
                        nb_lits_dispo INTEGER,
                        jour_ferie    INTEGER DEFAULT 0,
                        mutuelle      TEXT DEFAULT 'Payant',
                        prix_sejour   REAL DEFAULT 0,
                        prix_soins    REAL DEFAULT 0
                    )
                """))
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS urgences (
                        "ID_Urgence"        TEXT,
                        "ID_Patient"        TEXT,
                        "Nom_Complet"       TEXT,
                        "Age"               INTEGER,
                        "Sexe"              VARCHAR(5),
                        "CIN"               VARCHAR(20) DEFAULT '',
                        "Groupe_Sanguin"    VARCHAR(10),
                        "Antecedents"       TEXT,
                        "Etablissement"     TEXT,
                        "Type_Etab"         VARCHAR(50),
                        "Ville"             VARCHAR(100),
                        "Date_Arrivee"      TIMESTAMP,
                        "Date_Sortie"       TIMESTAMP,
                        "Niveau_Triage"     TEXT,
                        "Motif_Consultation" TEXT,
                        "Orientation"       TEXT,
                        "Duree_Sejour_min"  INTEGER,
                        "Nb_Medecins_Dispo" INTEGER,
                        "Nb_Lits_Dispo"     INTEGER,
                        "Jour_Ferie"        INTEGER,
                        "Saison"            TEXT,
                        "Heure_Arrivee"     INTEGER,
                        "Jour_Semaine"      INTEGER,
                        "Mois"              INTEGER,
                        "Annee"             INTEGER,
                        "Tranche_Horaire"   TEXT,
                        "Nom_Jour"          TEXT,
                        "Groupe_Age"        TEXT,
                        "Est_Pic"           INTEGER,
                        "Mutuelle"          VARCHAR(50),
                        "Prix_Sejour"       FLOAT,
                        "Prix_Soins"        FLOAT
                    )
                """))
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS soins_bronze (
                        id_soin       BIGSERIAL PRIMARY KEY,
                        id_urgence    BIGINT,
                        id_patient    BIGINT,
                        type_soin     TEXT,
                        description   TEXT,
                        cout          FLOAT,
                        date_soin     TIMESTAMP
                    )
                """))
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS soins (
                        "ID_Soin"       TEXT,
                        "ID_Urgence"    TEXT,
                        "ID_Patient"    TEXT,
                        "Type_Soin"     TEXT,
                        "Description"   TEXT,
                        "Cout"          FLOAT,
                        "Date_Soin"     TIMESTAMP
                    )
                """))
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS etablissements_bronze (
                        id            BIGSERIAL PRIMARY KEY,
                        nom           TEXT,
                        type_etab     TEXT,
                        ville         TEXT,
                        wilaya        TEXT,
                        capacite      INTEGER,
                        nb_medecins   INTEGER
                    )
                """))
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS etablissements (
                        "Nom"           TEXT,
                        "Type_Etab"     TEXT,
                        "Ville"         TEXT,
                        "Wilaya"        TEXT,
                        "Capacite"      INTEGER,
                        "Nb_Medecins"   INTEGER
                    )
                """))
                conn.commit()
            print("Tables PostgreSQL créées (ou déjà existantes).")

        # ── Charger en mémoire depuis PostgreSQL ou vide ──────────────
        reload_urg()

        if engine:
            try:
                data["soins"] = pd.read_sql('SELECT * FROM soins', engine)
            except Exception:
                data["soins"] = pd.DataFrame()
            try:
                data["etab"] = pd.read_sql('SELECT * FROM etablissements', engine)
            except Exception:
                data["etab"] = pd.DataFrame()
        else:
            data["soins"] = pd.DataFrame()
            data["etab"]  = pd.DataFrame()

        # Fichiers optionnels (prédictions / séries temporelles / métriques)
        for key, path in [
            ("ts",      os.path.join(GOLD,   "serie_temporelle_daily.csv")),
            ("pred30",  os.path.join(GOLD,   "predictions_30jours.csv")),
            ("metrics", os.path.join(MODELS, "metrics_comparison.csv")),
        ]:
            try:
                data[key] = pd.read_csv(path, encoding="utf-8-sig")
            except Exception:
                data[key] = pd.DataFrame()

        # Modèles ML optionnels
        for key, path in [
            ("xgb", os.path.join(MODELS, "xgboost_model.pkl")),
            ("rf",  os.path.join(MODELS, "random_forest_model.pkl")),
            ("le",  os.path.join(MODELS, "label_encoder.pkl")),
        ]:
            try:
                data[key] = joblib.load(path)
            except Exception:
                data[key] = None

        # Initialiser les tables temps réel
        _init_users_table()
        if engine:
            with engine.connect() as conn:
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS patient_statuts (
                        id_urgence TEXT PRIMARY KEY,
                        statut     TEXT NOT NULL,
                        lit_numero TEXT DEFAULT '',
                        updated_at TEXT,
                        updated_by TEXT
                    )
                """))
                # Migration : ajouter lit_numero si la colonne n'existe pas encore
                try:
                    conn.execute(text("ALTER TABLE patient_statuts ADD COLUMN lit_numero TEXT DEFAULT ''"))
                except Exception:
                    pass  # colonne déjà présente
                conn.commit()

        print("Données et modèles chargés avec succès")
    except Exception as e:
        print(f"ERREUR chargement : {e}")


def clean(val):
    """Convertit NaN/inf en None pour JSON."""
    if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
        return None
    return val

def df_to_records(df):
    return [
        {k: (None if isinstance(v, float) and (np.isnan(v) or np.isinf(v)) else
             round(v, 2) if isinstance(v, float) else v)
         for k, v in row.items()}
        for row in df.to_dict(orient="records")
    ]


# ══════════════════════════════════════════════════════════════
# KPIs
# ══════════════════════════════════════════════════════════════
@app.get("/api/kpis")
def get_kpis(annees: str = "", orientation: str = ""):
    try:
        urg = get_urg()
        if annees:
            annees_list = [int(a) for a in annees.split(",") if a.strip().isdigit()]
            if annees_list and "Annee" in urg.columns:
                urg = urg[urg["Annee"].isin(annees_list)]
        if orientation and orientation != "Toutes":
            urg = urg[urg["Orientation"] == orientation]
        total        = len(urg)
        if total == 0:
            return {"total": 0, "patients_par_jour": 0, "duree_moy": 0, "taux_hospit": 0, "taux_fugue": 0, "taux_p1": 0}
        nb_jours     = urg["Date_Arrivee"].dt.date.nunique()
        par_jour     = round(total / nb_jours, 1) if nb_jours > 0 else 0
        duree_moy    = round(urg["Duree_Sejour_min"].mean(), 1)
        taux_hospit  = round(len(urg[urg["Orientation"] == "Hospitalise"]) / total * 100, 2)
        taux_fugue   = round(len(urg[urg["Orientation"] == "Fugue"]) / total * 100, 2)
        taux_p1      = round(len(urg[urg["Niveau_Triage"].str.startswith("P1", na=False)]) / total * 100, 2)
        return {
            "total": total,
            "patients_par_jour": par_jour,
            "duree_moy": duree_moy,
            "taux_hospit": taux_hospit,
            "taux_fugue": taux_fugue,
            "taux_p1": taux_p1
        }
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# URGENCES
# ══════════════════════════════════════════════════════════════
@app.get("/api/urgences/temporel")
def get_temporel(annees: str = "2019,2020,2021,2022,2023,2024,2025,2026"):
    try:
        urg = get_urg()
        annees_list = [int(a) for a in annees.split(",") if a.strip().isdigit()]
        urg = urg[urg["Date_Arrivee"].dt.year.isin(annees_list)]
        ts = urg.groupby(urg["Date_Arrivee"].dt.date).size().reset_index()
        ts.columns = ["ds", "y"]
        ts["ds"] = ts["ds"].astype(str)
        ts = ts.sort_values("ds")
        return df_to_records(ts)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/urgences/horaire")
def get_horaire():
    try:
        urg = get_urg()
        col = urg["Heure_Arrivee"].dropna().astype(int)
        h = col.value_counts().sort_index().reset_index()
        h.columns = ["heure", "nb_patients"]
        return df_to_records(h)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/urgences/triage")
def get_triage():
    try:
        urg = get_urg().copy()
        triage_map = {
            "p1 - critique":      "P1 - Critique",
            "p1 - urgence absolue": "P1 - Critique",
            "critique":           "P1 - Critique",
            "immédiat":           "P1 - Critique",
            "immediat":           "P1 - Critique",
            "p2 - urgent":        "P2 - Urgent",
            "p2 - urgence relative": "P2 - Urgent",
            "urgent":             "P2 - Urgent",
            "très urgent":        "P2 - Urgent",
            "tres urgent":        "P2 - Urgent",
            "p3 - semi-urgent":   "P3 - Semi-urgent",
            "p3 - urgence différée": "P3 - Semi-urgent",
            "semi-urgent":        "P3 - Semi-urgent",
            "p4 - non urgent":    "P4 - Non urgent",
            "non urgent":         "P4 - Non urgent",
        }
        urg["Niveau_Triage"] = urg["Niveau_Triage"].str.strip().apply(
            lambda x: triage_map.get(str(x).lower(), str(x)) if pd.notna(x) else x
        )
        t = urg["Niveau_Triage"].value_counts().reset_index()
        t.columns = ["triage","count"]
        total = t["count"].sum()
        t["pct"] = (t["count"] / total * 100).round(2)
        return df_to_records(t)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/urgences/orientation")
def get_orientation():
    try:
        urg = get_urg().copy()
        # Normaliser les valeurs incohérentes
        orient_map = {
            "retour domicile": "Domicile", "retour_domicile": "Domicile",
            "domicile": "Domicile",
            "hospitalisation": "Hospitalise", "hospitalisé": "Hospitalise",
            "hospitalise": "Hospitalise",
            "transfert": "Transfere", "transferé": "Transfere",
            "transfere": "Transfere",
            "décès": "Decede", "deces": "Decede", "décédé": "Decede",
            "decede": "Decede",
            "fugue": "Fugue",
        }
        urg["Orientation"] = urg["Orientation"].str.strip().apply(
            lambda x: orient_map.get(str(x).lower(), str(x)) if pd.notna(x) else x
        )
        o = urg["Orientation"].value_counts().reset_index()
        o.columns = ["orientation","count"]
        total = o["count"].sum()
        o["pct"] = (o["count"] / total * 100).round(2)
        return df_to_records(o)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/urgences/annuel")
def get_annuel():
    try:
        urg = get_urg()
        a = urg.groupby("Annee").size().reset_index(name="nb_patients")
        a = a.sort_values("Annee")
        a["variation_pct"] = a["nb_patients"].pct_change().mul(100).round(2)
        a["variation_pct"] = a["variation_pct"].fillna(0)
        return df_to_records(a)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/urgences/saison")
def get_saison():
    try:
        urg = get_urg()
        ordre = ["Hiver","Printemps","Eté","Automne"]
        s = urg["Saison"].value_counts().reset_index()
        s.columns = ["saison","nb_patients"]
        s["saison"] = s["saison"].str.replace("Été","Eté")
        s["order"] = s["saison"].apply(lambda x: ordre.index(x) if x in ordre else 99)
        s = s.sort_values("order").drop(columns="order")
        return df_to_records(s)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/urgences/jour")
def get_jour():
    try:
        urg = get_urg()
        jours = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]
        col = urg["Jour_Semaine"].dropna().astype(int)
        j = col.value_counts().sort_index().reset_index()
        j.columns = ["Jour_Semaine", "nb_patients"]
        j["jour"] = j["Jour_Semaine"].apply(lambda x: jours[x] if 0 <= x <= 6 else str(x))
        return df_to_records(j[["jour", "nb_patients"]])
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/urgences/top_motifs")
def get_top_motifs():
    try:
        urg = get_urg().copy()
        if "Motif_Consultation" not in urg.columns:
            return []
        total = len(urg)
        motif_counts = urg["Motif_Consultation"].value_counts()
        results = []
        for motif, count in motif_counts.head(15).items():
            grp = urg[urg["Motif_Consultation"] == motif]
            # triage le plus fréquent
            triage_dist = grp["Niveau_Triage"].value_counts()
            triage_principal = str(triage_dist.index[0]) if len(triage_dist) else "N/A"
            # taux P1
            p1_rate = round(float((grp["Niveau_Triage"] == "P1 - Critique").sum()) / len(grp) * 100, 1)
            # taux hospitalisation
            hospit_rate = round(float((grp["Orientation"] == "Hospitalise").sum()) / len(grp) * 100, 1)
            # durée moyenne
            avg_duree = int(grp["Duree_Sejour_min"].mean()) if "Duree_Sejour_min" in grp.columns else 0
            # tendance : comparer dernière année vs avant
            if "Annee" in urg.columns:
                annee_max = int(urg["Annee"].max())
                count_recent = int((grp["Annee"] == annee_max).sum())
                count_avant  = int((grp["Annee"] == annee_max - 1).sum())
                if count_avant > 0:
                    tendance = round((count_recent - count_avant) / count_avant * 100, 1)
                else:
                    tendance = 0.0
            else:
                tendance = 0.0
            results.append({
                "motif":             str(motif),
                "count":             int(count),
                "pct":               round(count / total * 100, 1),
                "triage_principal":  triage_principal,
                "p1_rate":           p1_rate,
                "hospit_rate":       hospit_rate,
                "avg_duree":         avg_duree,
                "tendance":          tendance,
            })
        return results
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/urgences/maladies_saisonnieres")
def get_maladies_saisonnieres():
    """
    Pour chaque saison, retourne le top 5 des motifs de consultation
    avec le nombre de cas, le taux de criticité P1 et la durée moyenne.
    """
    try:
        urg = get_urg().copy()
        if "Motif_Consultation" not in urg.columns or "Saison" not in urg.columns:
            return []

        # Normaliser toutes les variantes d'Été → "Ete" (valeur réelle en DB)
        urg["Saison"] = urg["Saison"].str.replace("Été", "Ete").str.replace("Eté", "Ete")
        saisons_ordre = ["Hiver", "Printemps", "Ete", "Automne"]

        result = []
        for saison in saisons_ordre:
            grp_saison = urg[urg["Saison"] == saison]
            if grp_saison.empty:
                continue
            total_saison = len(grp_saison)
            top_motifs = grp_saison["Motif_Consultation"].value_counts().head(5)
            maladies = []
            for motif, count in top_motifs.items():
                grp_m = grp_saison[grp_saison["Motif_Consultation"] == motif]
                p1 = round(float((grp_m["Niveau_Triage"] == "P1 - Critique").sum()) / len(grp_m) * 100, 1)
                duree = int(grp_m["Duree_Sejour_min"].mean()) if "Duree_Sejour_min" in grp_m.columns else 0
                hospit = round(float((grp_m["Orientation"] == "Hospitalise").sum()) / len(grp_m) * 100, 1)
                maladies.append({
                    "motif": str(motif),
                    "count": int(count),
                    "pct": round(count / total_saison * 100, 1),
                    "p1_rate": p1,
                    "hospit_rate": hospit,
                    "avg_duree": duree,
                })
            result.append({
                "saison": saison,
                "total_patients": total_saison,
                "maladies": maladies,
            })
        return result
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# ETABLISSEMENTS
# ══════════════════════════════════════════════════════════════
@app.get("/api/etablissements")
def get_etablissements():
    try:
        return df_to_records(data["etab"])
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/ressources")
def get_ressources():
    try:
        urg  = get_urg().copy()
        etab = data.get("etab", pd.DataFrame())

        # Build a lookup from the real etablissements table (capacite_lits, nb_medecins, nb_urgentistes)
        etab_lookup: dict = {}
        if not etab.empty and "nom" in etab.columns:
            for _, row in etab.iterrows():
                etab_lookup[row["nom"]] = {
                    "medecins":   int(row.get("nb_medecins",   20)),
                    "infirmiers": int(row.get("nb_urgentistes", 40)),
                    "lits_total": int(row.get("capacite_lits",  50)),
                }

        # Fallback defaults for hospitals not in the etablissements table
        FALLBACK = {
            "CHU Ibn Sina":    {"medecins": 60, "infirmiers": 120, "lits_total": 200},
            "Hopital Ibn Sina":{"medecins": 45, "infirmiers":  90, "lits_total": 120},
            "Hopital des Enfants": {"medecins": 25, "infirmiers": 50, "lits_total": 60},
            "Hopital Al Ayachi":   {"medecins": 30, "infirmiers": 60, "lits_total": 80},
            "Hopital Ar-Razi":     {"medecins": 20, "infirmiers": 40, "lits_total": 50},
            "Hopital des Specialites": {"medecins": 35, "infirmiers": 70, "lits_total": 100},
            "Hopital de Maternite et de Sante Reproductrice les Orangers": {"medecins": 25, "infirmiers": 50, "lits_total": 80},
            "Hopital Moulay Youssef": {"medecins": 20, "infirmiers": 40, "lits_total": 60},
            "Hopital de Maternite Souissi": {"medecins": 20, "infirmiers": 40, "lits_total": 60},
        }

        results = []
        etablissements = urg["Etablissement"].dropna().unique() if "Etablissement" in urg.columns else []

        for etab_nom in etablissements:
            grp     = urg[urg["Etablissement"] == etab_nom]
            # Use real table data → fallback constants → generic default
            ressrc  = etab_lookup.get(str(etab_nom)) or FALLBACK.get(str(etab_nom)) or {"medecins": 20, "infirmiers": 40, "lits_total": 50}

            # Lits occupés = patients hospitalisés sur les 30 derniers jours
            recents      = grp[grp["Date_Arrivee"] >= pd.Timestamp.now() - pd.Timedelta(days=30)]
            lits_occupes = int((recents["Orientation"] == "Hospitalise").sum()) if len(recents) > 0 else 0
            lits_occupes = min(lits_occupes, ressrc["lits_total"])
            lits_dispo   = ressrc["lits_total"] - lits_occupes
            taux_occup   = round(lits_occupes / ressrc["lits_total"] * 100, 1) if ressrc["lits_total"] > 0 else 0
            charge       = "Critique" if taux_occup >= 80 else "Élevée" if taux_occup >= 60 else "Normale"

            results.append({
                "etablissement":   str(etab_nom),
                "medecins":        ressrc["medecins"],
                "infirmiers":      ressrc["infirmiers"],
                "lits_total":      ressrc["lits_total"],
                "lits_occupes":    lits_occupes,
                "lits_dispo":      lits_dispo,
                "taux_occupation": taux_occup,
                "charge":          charge,
                "total_patients":  int(len(grp)),
            })

        results.sort(key=lambda x: x["taux_occupation"], reverse=True)
        return results
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# SOINS
# ══════════════════════════════════════════════════════════════
@app.get("/api/soins/types")
def get_soins_types():
    try:
        s = get_soins()["Type_Soin"].value_counts().reset_index()
        s.columns = ["type_soin","count"]
        return df_to_records(s.head(15))
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/soins/couts_par_type")
def get_couts_par_type():
    try:
        s = get_soins().groupby("Type_Soin")["Cout_Soin"].mean().reset_index()
        s.columns = ["type_soin","cout_moyen"]
        s["cout_moyen"] = s["cout_moyen"].round(2)
        s = s.sort_values("cout_moyen", ascending=False).head(15)
        return df_to_records(s)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/soins/couts_par_etab")
def get_couts_par_etab():
    try:
        s = get_soins().groupby("Etablissement")["Cout_Soin"].sum().reset_index()
        s.columns = ["etablissement","cout_total"]
        s["cout_total"] = s["cout_total"].round(2)
        s = s.sort_values("cout_total", ascending=False)
        return df_to_records(s)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/soins/resultats")
def get_resultats():
    try:
        s = get_soins()["Resultat"].value_counts().reset_index()
        s.columns = ["resultat","count"]
        return df_to_records(s)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/soins/medicaments")
def get_medicaments():
    try:
        s = get_soins()["Medicament"].value_counts().reset_index()
        s.columns = ["medicament","count"]
        return df_to_records(s.head(12))
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# PREDICTIONS
# ══════════════════════════════════════════════════════════════
@app.get("/api/predictions")
def get_predictions():
    try:
        # Si pred30 CSV disponible, l'utiliser
        if not data["pred30"].empty:
            return df_to_records(data["pred30"])
        # Sinon : simulation basique 30j depuis la moyenne historique par jour
        urg = get_urg()
        daily = urg.groupby(urg["Date_Arrivee"].dt.date).size().reset_index()
        daily.columns = ["ds", "y"]
        daily["ds"] = pd.to_datetime(daily["ds"])
        avg = daily["y"].mean()
        std = daily["y"].std()
        last_date = daily["ds"].max()
        future = pd.date_range(start=last_date + pd.Timedelta(days=1), periods=30)
        import numpy as np
        pred = pd.DataFrame({
            "ds":    future.strftime("%Y-%m-%d"),
            "yhat":  np.round(np.random.normal(avg, std * 0.3, 30).clip(min=1)).astype(int),
            "yhat_lower": np.round((avg - std * 0.5)).astype(int),
            "yhat_upper": np.round((avg + std * 0.5)).astype(int),
        })
        return df_to_records(pred)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/modeles/metriques")
def get_metriques():
    try:
        return df_to_records(data["metrics"])
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# SIMULATEUR XGBoost
# ══════════════════════════════════════════════════════════════
class SimulateurInput(BaseModel):
    mois: int
    heure: int
    jour_semaine: int
    jour_ferie: int
    saison: str

SAISON_MAP = {"Hiver": 0, "Printemps": 1, "Eté": 2, "Ete": 2, "Automne": 3}

@app.post("/api/simulateur")
def simulateur(body: SimulateurInput):
    try:
        urg = get_urg()
        xgb = data["xgb"]

        saison_enc   = SAISON_MAP.get(body.saison, 0)
        est_weekend  = 1 if body.jour_semaine >= 5 else 0
        trimestre    = (body.mois - 1) // 3 + 1

        # Moyennes historiques
        nb_medecins  = float(urg["Nb_Medecins_Dispo"].mean())
        nb_lits      = float(urg["Nb_Lits_Dispo"].mean())
        lits_par_med = nb_lits / (nb_medecins + 1)

        # Encodage triage moyen (P2)
        triage_enc   = 3
        orient_enc   = 1
        age_moy      = float(urg["Age"].mean())
        etab_moy     = 3  # milieu
        groupe_age   = 3
        est_pic      = 1 if 8 <= body.heure <= 12 or 17 <= body.heure <= 21 else 0
        annee        = 2024

        heure_sin = np.sin(2*np.pi*body.heure/24)
        heure_cos = np.cos(2*np.pi*body.heure/24)
        jour_sin  = np.sin(2*np.pi*body.jour_semaine/7)
        jour_cos  = np.cos(2*np.pi*body.jour_semaine/7)
        mois_sin  = np.sin(2*np.pi*body.mois/12)
        mois_cos  = np.cos(2*np.pi*body.mois/12)

        X = pd.DataFrame([{
            "Age": age_moy,
            "sexe_enc": 0,
            "triage_enc": triage_enc,
            "orientation_enc": orient_enc,
            "ferie_enc": body.jour_ferie,
            "weekend": est_weekend,
            "saison_enc": saison_enc,
            "etab_enc": etab_moy,
            "groupe_age_enc": groupe_age,
            "Heure_Arrivee": body.heure,
            "Jour_Semaine": body.jour_semaine,
            "Mois": body.mois,
            "Annee": annee,
            "Nb_Medecins_Dispo": nb_medecins,
            "Nb_Lits_Dispo": nb_lits,
            "lits_par_med": lits_par_med,
            "heure_sin": heure_sin,
            "heure_cos": heure_cos,
            "jour_sin": jour_sin,
            "jour_cos": jour_cos,
            "mois_sin": mois_sin,
            "mois_cos": mois_cos,
            "triage_x_age": triage_enc * age_moy,
            "triage_x_etab": triage_enc * etab_moy,
            "Est_Pic": est_pic,
        }])

        pred = float(xgb.predict(X)[0])
        pred = round(max(0, pred), 1)

        if pred < 150:
            niveau = "FAIBLE"
            couleur = "#22C55E"
        elif pred < 250:
            niveau = "MODERE"
            couleur = "#F59E0B"
        else:
            niveau = "ELEVE"
            couleur = "#EF4444"

        return {"prediction": pred, "niveau": niveau, "couleur": couleur}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/")
def root():
    return {"message": "CHU Ibn Sina API", "docs": "/docs"}


# ══════════════════════════════════════════════════════════════
# AUTH — JWT
# ══════════════════════════════════════════════════════════════
JWT_SECRET  = os.environ.get("JWT_SECRET", "chu_ibn_sina_secret_2025_pfe")
JWT_ALGO    = "HS256"
JWT_EXPIRE  = int(os.environ.get("JWT_EXPIRE_MINUTES", 480))

def _hash(pwd: str) -> str:
    return hashlib.sha256(pwd.encode()).hexdigest()

USERS = {
    "admin":        {"password": _hash("admin123"), "role": "admin"},
    "chef_medecin": {"password": _hash("admin123"), "role": "chef_medecin"},
    "urgentiste":   {"password": _hash("admin123"), "role": "urgentiste"},
    "infirmier":    {"password": _hash("admin123"), "role": "infirmier"},
    "directeur":    {"password": _hash("admin123"), "role": "directeur"},
    "analyste":     {"password": _hash("admin123"), "role": "analyste"},
    "admin_si":     {"password": _hash("admin123"), "role": "admin_si"},
}

security = HTTPBearer(auto_error=False)

def create_jwt(username: str, role: str, etablissement: str = "") -> str:
    if JWT_AVAILABLE:
        payload = {
            "sub":          username,
            "role":         role,
            "etablissement": etablissement,
            "exp":          datetime.utcnow() + timedelta(minutes=JWT_EXPIRE),
            "iat":          datetime.utcnow(),
        }
        return pyjwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)
    else:
        import secrets as _s
        return _s.token_hex(32)

def decode_jwt(token: str) -> Optional[dict]:
    if JWT_AVAILABLE:
        try:
            return pyjwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
        except Exception:
            return None
    return None

class LoginInput(BaseModel):
    username: str
    password: str

@app.post("/api/auth/logout")
def logout():
    # JWT est stateless — le client supprime simplement le token
    return {"message": "Deconnecte"}

@app.get("/api/auth/me")
def me(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if not credentials:
        raise HTTPException(status_code=401, detail="Non authentifie")
    payload = decode_jwt(credentials.credentials)
    if not payload:
        raise HTTPException(status_code=401, detail="Token invalide ou expire")
    return {"username": payload.get("sub"), "role": payload.get("role")}

@app.post("/api/auth/refresh")
def refresh(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Renouvelle le JWT si encore valide."""
    if not credentials:
        raise HTTPException(status_code=401, detail="Non authentifie")
    payload = decode_jwt(credentials.credentials)
    if not payload:
        raise HTTPException(status_code=401, detail="Token invalide ou expire")
    username = payload.get("sub")
    role = payload.get("role")
    new_token = create_jwt(username, role)
    return {"token": new_token, "username": username, "role": role, "expires_in": JWT_EXPIRE * 60}


# ══════════════════════════════════════════════════════════════
# ALERTES
# ══════════════════════════════════════════════════════════════
@app.get("/api/alertes")
def get_alertes():
    try:
        etab = data["etab"]
        alertes = []
        for _, row in etab.iterrows():
            if row.get("Taux_Hospit_Pct", 0) > 35:
                alertes.append({
                    "etablissement": row["nom"],
                    "type": "Taux hospitalisation élevé",
                    "valeur": f"{row['Taux_Hospit_Pct']}%",
                    "seuil": "35%",
                    "niveau": "critique" if row["Taux_Hospit_Pct"] > 45 else "warning",
                })
            if row.get("Taux_Fugue_Pct", 0) > 12:
                alertes.append({
                    "etablissement": row["nom"],
                    "type": "Taux de fugue élevé",
                    "valeur": f"{row['Taux_Fugue_Pct']}%",
                    "seuil": "12%",
                    "niveau": "warning",
                })
            if row.get("Duree_Moy_Min", 0) > 300:
                alertes.append({
                    "etablissement": row["nom"],
                    "type": "Durée séjour excessive",
                    "valeur": f"{row['Duree_Moy_Min']} min",
                    "seuil": "300 min",
                    "niveau": "critique",
                })
        return alertes
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# STATS RESUME
# ══════════════════════════════════════════════════════════════
@app.get("/api/stats/resume")
def get_stats_resume():
    try:
        urg = get_urg()
        s = urg["Duree_Sejour_min"]
        return {
            "duree_sejour": {
                "min":    round(float(s.min()), 1),
                "max":    round(float(s.max()), 1),
                "mean":   round(float(s.mean()), 1),
                "median": round(float(s.median()), 1),
                "p25":    round(float(s.quantile(0.25)), 1),
                "p75":    round(float(s.quantile(0.75)), 1),
                "p90":    round(float(s.quantile(0.90)), 1),
                "std":    round(float(s.std()), 1),
            },
            "age": {
                "mean":   round(float(urg["Age"].mean()), 1),
                "median": round(float(urg["Age"].median()), 1),
                "min":    int(urg["Age"].min()),
                "max":    int(urg["Age"].max()),
            },
            "triage_dist": urg["Niveau_Triage"].value_counts().to_dict(),
            "orientation_dist": urg["Orientation"].value_counts().to_dict(),
            "saison_dist": urg["Saison"].value_counts().to_dict(),
            "groupe_age_dist": urg["Groupe_Age"].value_counts().to_dict(),
            "annees": sorted(urg["Annee"].dropna().astype(int).unique().tolist()),
            "total": len(urg),
        }
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# EXPORT CSV
# ══════════════════════════════════════════════════════════════
@app.get("/api/export/urgences")
def export_urgences():
    try:
        buf = io.StringIO()
        get_urg().to_csv(buf, index=False, encoding="utf-8-sig")
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=urgences_CHU_Ibn_Sina.csv"}
        )
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/export/soins")
def export_soins():
    try:
        buf = io.StringIO()
        get_soins().to_csv(buf, index=False, encoding="utf-8-sig")
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=soins_CHU_Ibn_Sina.csv"}
        )
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/export/etablissements")
def export_etablissements():
    try:
        buf = io.StringIO()
        data["etab"].to_csv(buf, index=False, encoding="utf-8-sig")
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=etablissements_CHU_Ibn_Sina.csv"}
        )
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# ETABLISSEMENTS — Coordonnées géographiques
# ══════════════════════════════════════════════════════════════
COORDS = {
    "Hopital Ibn Sina":   {"lat": 34.0209, "lng": -6.8416},
    "Hopital des Enfants": {"lat": 34.0132, "lng": -6.8326},
    "Hopital Al Ayachi":  {"lat": 34.0369, "lng": -6.8326},
    "Hopital Ar-Razi":    {"lat": 34.0442, "lng": -6.7985},
    "Hopital des Specialites": {"lat": 34.0178, "lng": -6.8356},
    "Hopital de Maternite et de Sante Reproductrice les Orangers": {"lat": 34.0089, "lng": -6.8512},
    "Hopital Moulay Youssef": {"lat": 34.0156, "lng": -6.8423},
    "Hopital de Maternite Souissi": {"lat": 33.9956, "lng": -6.8512},
}

@app.get("/api/urgences/liste")
def get_urgences_liste(limit: int = 20000):
    try:
        urg = get_urg()
        cols = ["ID_Urgence", "Nom_Complet", "CIN", "Age", "Sexe",
                "Groupe_Sanguin", "Antecedents", "Niveau_Triage", "Date_Arrivee",
                "Etablissement", "Orientation", "Duree_Sejour_min", "Saison", "Annee",
                "Mutuelle", "Prix_Sejour", "Prix_Soins"]
        available = [c for c in cols if c in urg.columns]
        df = urg[available].copy()
        if "Date_Arrivee" in df.columns:
            df = df.sort_values("Date_Arrivee", ascending=False)
        df = df.head(limit)
        df["Date_Arrivee"] = df["Date_Arrivee"].astype(str)
        if "ID_Urgence" in df.columns:
            df = df.rename(columns={"ID_Urgence": "id_passage"})
        return df_to_records(df)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/patients/liste")
def get_patients_liste():
    """Vue patient : une ligne par patient avec tout son historique."""
    try:
        urg = get_urg().copy()
        urg["Date_Arrivee"] = pd.to_datetime(urg["Date_Arrivee"], errors="coerce")
        urg = urg.sort_values("Date_Arrivee", ascending=False).head(20000)

        TRIAGE_ORDER = {"P1 - Critique": 1, "P2 - Urgent": 2, "P3 - Semi-urgent": 3, "P4 - Non urgent": 4}

        results = []
        for nom, grp in urg.groupby("Nom_Complet", sort=False):
            grp = grp.sort_values("Date_Arrivee", ascending=False)
            derniere = grp.iloc[0]
            niveaux = grp["Niveau_Triage"].dropna().tolist()
            pire = min(niveaux, key=lambda x: TRIAGE_ORDER.get(x, 99)) if niveaux else ""
            historique = []
            for _, row in grp.iterrows():
                historique.append({
                    "id_passage":    str(row.get("ID_Urgence", "")),
                    "date":          str(row["Date_Arrivee"])[:10] if pd.notna(row["Date_Arrivee"]) else "",
                    "niveau_triage": str(row.get("Niveau_Triage", "")),
                    "motif":         str(row.get("Motif_Consultation", "")),
                    "orientation":   str(row.get("Orientation", "")),
                    "duree_min":     int(row["Duree_Sejour_min"]) if pd.notna(row.get("Duree_Sejour_min")) else 0,
                    "etablissement": str(row.get("Etablissement", "")),
                    "mutuelle":      str(row.get("Mutuelle", "Payant")),
                    "prix_sejour":   float(row["Prix_Sejour"]) if pd.notna(row.get("Prix_Sejour")) else 0.0,
                    "prix_soins":    float(row["Prix_Soins"])  if pd.notna(row.get("Prix_Soins"))  else 0.0,
                })
            total_sejour = float(grp["Prix_Sejour"].sum()) if "Prix_Sejour" in grp.columns else 0.0
            total_soins  = float(grp["Prix_Soins"].sum())  if "Prix_Soins"  in grp.columns else 0.0
            results.append({
                "nom_complet":      nom,
                "age":              int(derniere["Age"]) if pd.notna(derniere.get("Age")) else 0,
                "sexe":             str(derniere.get("Sexe", "")),
                "cin":              str(derniere.get("CIN", "")),
                "groupe_sanguin":   str(derniere.get("Groupe_Sanguin", "")),
                "antecedents":      str(derniere.get("Antecedents", "")),
                "mutuelle":         str(derniere.get("Mutuelle", "Payant")),
                "total_sejour":     total_sejour,
                "total_soins":      total_soins,
                "nb_visites":       len(grp),
                "derniere_visite":  str(derniere["Date_Arrivee"])[:10] if pd.notna(derniere["Date_Arrivee"]) else "",
                "triage_max":       pire,
                "historique":       historique,
            })
        results.sort(key=lambda x: x["derniere_visite"], reverse=True)
        return results
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/etablissements/carte")
def get_carte():
    try:
        etab = data["etab"]
        result = []
        for _, row in etab.iterrows():
            coords = COORDS.get(row["nom"], {"lat": 34.02, "lng": -6.84})
            result.append({
                "nom": row["nom"],
                "type_etab": row["type_etab"],
                "ville": row["ville"],
                "lat": coords["lat"],
                "lng": coords["lng"],
                "capacite_lits": int(row["capacite_lits"]),
                "nb_medecins": int(row["nb_medecins"]),
                "Nb_Patients": int(row.get("Nb_Patients", 0)),
                "Taux_Hospit_Pct": round(float(row.get("Taux_Hospit_Pct", 0)), 2),
                "Duree_Moy_Min": round(float(row.get("Duree_Moy_Min", 0)), 1),
                "Alerte_Charge": row.get("Alerte_Charge", "Normal"),
            })
        return result
    except Exception as e:
        raise HTTPException(500, str(e))



# ══════════════════════════════════════════════════════════════
# PRÉDICTION TRIAGE À L'ADMISSION
# ══════════════════════════════════════════════════════════════
class TriageInput(BaseModel):
    age: int
    sexe: str           # "M" ou "F"
    heure: int          # 0-23
    jour_semaine: int   # 0=Lundi ... 6=Dimanche
    mois: int           # 1-12
    saison: str         # "Hiver","Printemps","Été","Automne"
    jour_ferie: bool = False
    antecedents: str = "Aucun"

TRIAGE_LABELS = {1: "P1 - Critique", 2: "P2 - Urgent", 3: "P3 - Semi-urgent", 4: "P4 - Non urgent"}
TRIAGE_COLORS = {1: "#EF4444", 2: "#F59E0B", 3: "#3B82F6", 4: "#22C55E"}
TRIAGE_RISK   = {1: "Critique", 2: "Élevé", 3: "Modéré", 4: "Faible"}

SAISON_MAP = {"Hiver": 0, "Printemps": 1, "Été": 2, "Automne": 3}
SEXE_MAP   = {"M": 1, "F": 0}

def has_antecedent(ants: str, keyword: str) -> int:
    return int(keyword.lower() in ants.lower())

@app.post("/api/predict/triage")
def predict_triage(body: TriageInput):
    try:
        import math
        heure_sin = math.sin(2 * math.pi * body.heure / 24)
        heure_cos = math.cos(2 * math.pi * body.heure / 24)
        jour_sin  = math.sin(2 * math.pi * body.jour_semaine / 7)
        jour_cos  = math.cos(2 * math.pi * body.jour_semaine / 7)
        mois_sin  = math.sin(2 * math.pi * body.mois / 12)
        mois_cos  = math.cos(2 * math.pi * body.mois / 12)
        saison_n  = SAISON_MAP.get(body.saison, 0)
        sexe_n    = SEXE_MAP.get(body.sexe, 0)
        est_weekend = int(body.jour_semaine >= 5)

        # Estimation simple basée sur les facteurs de risque
        # (le modèle XGBoost prédit la durée, pas le triage — on utilise une règle experte)
        score = 0
        if body.age >= 70 or body.age <= 2: score += 2
        elif body.age >= 60: score += 1
        if body.heure in range(22, 24) or body.heure in range(0, 6): score += 1
        if body.jour_ferie: score += 1
        if has_antecedent(body.antecedents, "cardiaque"): score += 2
        if has_antecedent(body.antecedents, "diabète"): score += 1
        if has_antecedent(body.antecedents, "respiratoire"): score += 2
        if has_antecedent(body.antecedents, "neurologique"): score += 2
        if has_antecedent(body.antecedents, "cancer"): score += 2
        if body.saison == "Hiver": score += 1

        # Convertir score en niveau de triage
        if score >= 5: triage_num = 1
        elif score >= 3: triage_num = 2
        elif score >= 1: triage_num = 3
        else: triage_num = 4

        # Estimation durée via XGBoost si disponible
        duree_estimee = None
        if "xgb" in data:
            try:
                feat = pd.DataFrame([{
                    "Age": body.age, "Sexe_num": sexe_n,
                    "Triage_num": triage_num, "Orientation_num": 1,
                    "Heure": body.heure, "Jour_Semaine": body.jour_semaine,
                    "Mois": body.mois, "Annee": 2025,
                    "Saison_num": saison_n,
                    "Nb_Medecins_Dispo": 8, "Nb_Lits_Dispo": 30,
                    "Jour_Ferie": int(body.jour_ferie),
                    "Est_Pic": 0, "Est_Weekend": est_weekend,
                    "Heure_sin": heure_sin, "Heure_cos": heure_cos,
                    "Jour_sin": jour_sin, "Jour_cos": jour_cos,
                    "Mois_sin": mois_sin, "Mois_cos": mois_cos,
                    "Triage_Age_inter": triage_num * body.age,
                    "Triage_Heure_inter": triage_num * body.heure,
                    "Age_sq": body.age ** 2,
                    "Triage_Weekend": triage_num * est_weekend,
                }])
                duree_estimee = round(float(data["xgb"].predict(feat)[0]), 0)
            except Exception:
                pass

        # ── Prédiction des prix ────────────────────────────────────
        BASE_SEJOUR = {1: 4000, 2: 2500, 3: 1200, 4: 500}
        BASE_SOINS  = {1: 3000, 2: 1800, 3: 900,  4: 300}

        prix_sejour = BASE_SEJOUR[triage_num]
        prix_soins  = BASE_SOINS[triage_num]

        # Facteurs multiplicateurs
        if body.heure in range(22, 24) or body.heure in range(0, 6):
            prix_sejour = round(prix_sejour * 1.2)   # Majoration nuit +20%
            prix_soins  = round(prix_soins  * 1.2)
        if body.jour_ferie:
            prix_sejour = round(prix_sejour * 1.15)  # Majoration férié +15%
            prix_soins  = round(prix_soins  * 1.15)
        if has_antecedent(body.antecedents, "cardiaque") or \
           has_antecedent(body.antecedents, "cancer")    or \
           has_antecedent(body.antecedents, "neurologique"):
            prix_soins = round(prix_soins * 1.4)     # Antécédents graves +40%
        if has_antecedent(body.antecedents, "diabète") or \
           has_antecedent(body.antecedents, "respiratoire"):
            prix_soins = round(prix_soins * 1.2)     # Antécédents modérés +20%
        if duree_estimee and duree_estimee > 300:
            prix_sejour = round(prix_sejour * 1.3)   # Long séjour +30%
        if body.age >= 70 or body.age <= 2:
            prix_soins = round(prix_soins * 1.15)    # Âge extrême +15%

        prix_total = prix_sejour + prix_soins

        return {
            "triage": TRIAGE_LABELS[triage_num],
            "triage_num": triage_num,
            "color": TRIAGE_COLORS[triage_num],
            "risque": TRIAGE_RISK[triage_num],
            "score": score,
            "duree_estimee_min": duree_estimee,
            "facteurs": {
                "age_risque": body.age >= 60 or body.age <= 2,
                "heure_nuit": body.heure in range(22, 24) or body.heure in range(0, 6),
                "jour_ferie": body.jour_ferie,
                "antecedents_graves": any(
                    has_antecedent(body.antecedents, k)
                    for k in ["cardiaque", "respiratoire", "neurologique", "cancer"]
                ),
            },
            "prix": {
                "sejour":  prix_sejour,
                "soins":   prix_soins,
                "total":   prix_total,
            }
        }
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# HEATMAP — Passages par heure × jour de semaine
# ══════════════════════════════════════════════════════════════
@app.get("/api/heatmap")
def get_heatmap():
    try:
        urg = get_urg()
        JOURS = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
        # Jour_Semaine : 0=Lundi..6=Dimanche (pandas weekday)
        urg2 = urg.copy()
        urg2["jour_num"] = pd.to_datetime(urg2["Date_Arrivee"]).dt.weekday
        hm = urg2.groupby(["Heure_Arrivee", "jour_num"]).size().reset_index(name="count")
        result = []
        for _, row in hm.iterrows():
            result.append({
                "heure": int(row["Heure_Arrivee"]),
                "jour":  JOURS[int(row["jour_num"])],
                "jour_num": int(row["jour_num"]),
                "count": int(row["count"]),
            })
        return result
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# EXPORT EXCEL
# ══════════════════════════════════════════════════════════════
@app.get("/api/export/excel")
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            get_urg().to_excel(writer, sheet_name="Urgences", index=False)
            get_soins().to_excel(writer, sheet_name="Soins", index=False)
            data["etab"].to_excel(writer, sheet_name="Etablissements", index=False)
            # Style header row for each sheet
            wb = writer.book
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1A3BDB")
                    cell.alignment = Alignment(horizontal="center")
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=8)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=CHU_Ibn_Sina_Data.xlsx"}
        )
    except ImportError:
        raise HTTPException(500, detail="openpyxl non installé. Ajoutez-le à requirements.txt.")
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# STATS COMPARAISON PÉRIODES (semaine actuelle vs précédente)
# ══════════════════════════════════════════════════════════════
@app.get("/api/stats/comparaison")
def get_comparaison_periodes():
    try:
        urg = get_urg()
        urg2 = urg.copy()
        urg2["date"] = pd.to_datetime(urg2["Date_Arrivee"])
        max_date = urg2["date"].max()
        # Semaine actuelle = 7 derniers jours du dataset
        sem_act  = urg2[urg2["date"] > max_date - pd.Timedelta(days=7)]
        sem_prec = urg2[(urg2["date"] > max_date - pd.Timedelta(days=14)) &
                        (urg2["date"] <= max_date - pd.Timedelta(days=7))]
        def stats(df):
            if len(df) == 0:
                return {"patients": 0, "duree_moy": 0, "taux_fugue": 0, "taux_p1": 0}
            return {
                "patients":    len(df),
                "duree_moy":   round(float(df["Duree_Sejour_min"].mean()), 1),
                "taux_fugue":  round(len(df[df["Orientation"]=="Fugue"])/len(df)*100, 2),
                "taux_p1":     round(len(df[df["Niveau_Triage"].str.startswith("P1",na=False)])/len(df)*100, 2),
            }
        act  = stats(sem_act)
        prec = stats(sem_prec)
        def delta(a, b):
            return round(a - b, 2) if b != 0 else 0
        return {
            "actuelle":   act,
            "precedente": prec,
            "delta": {k: delta(act[k], prec[k]) for k in act},
            "periode_actuelle":   max_date.strftime("%d/%m/%Y"),
            "periode_precedente": (max_date - pd.Timedelta(days=7)).strftime("%d/%m/%Y"),
        }
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# ADMISSION PATIENT — Insertion nouvelle entrée urgence
# ══════════════════════════════════════════════════════════════
class AdmissionBody(BaseModel):
    nom_complet: str
    cin: str = ''
    age: int
    sexe: str                  # M / F
    groupe_sanguin: str
    antecedents: str
    etablissement: str
    niveau_triage: str
    motif_consultation: str
    orientation: str
    duree_sejour_min: int
    nb_medecins_dispo: int = 5
    nb_lits_dispo: int = 10
    date_arrivee: Optional[str] = None   # ISO string, défaut = maintenant
    mutuelle: str = 'Payant'
    prix_sejour: float = 0.0
    prix_soins: float = 0.0

@app.post("/api/patients/add")
def add_patient(body: AdmissionBody, creds: HTTPAuthorizationCredentials = Depends(security)):
    try:
        import traceback
        urg_path = os.path.join(GOLD, "urgences_GOLD.csv")
        df = get_urg()

        # Date / heure
        try:
            dt = pd.to_datetime(body.date_arrivee) if body.date_arrivee else pd.Timestamp.now()
        except Exception:
            dt = pd.Timestamp.now()

        date_sortie = dt + pd.Timedelta(minutes=body.duree_sejour_min)

        # IDs auto-incrémentés — compatible SQLite et PostgreSQL
        try:
            with engine.connect() as conn:
                val = conn.execute(text('SELECT MAX(CAST("ID_Urgence" AS INTEGER)) FROM urgences')).scalar()
            max_id_urg = int(val or 0) + 1
        except Exception:
            max_id_urg = int(df.shape[0]) + 1
        try:
            with engine.connect() as conn:
                val = conn.execute(text('SELECT MAX(CAST("ID_Patient" AS INTEGER)) FROM urgences')).scalar()
            max_id_pat = int(val or 0) + 1
        except Exception:
            max_id_pat = int(df.shape[0]) + 1

        SAISONS = {12:"Hiver",1:"Hiver",2:"Hiver",3:"Printemps",4:"Printemps",5:"Printemps",
                   6:"Ete",7:"Ete",8:"Ete",9:"Automne",10:"Automne",11:"Automne"}
        JOURS_FR = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]
        TRANCHES = {0:"Nuit",1:"Nuit",2:"Nuit",3:"Nuit",4:"Nuit",5:"Nuit",
                    6:"Matin",7:"Matin",8:"Matin",9:"Matin",10:"Matin",11:"Matin",
                    12:"Apres-midi",13:"Apres-midi",14:"Apres-midi",15:"Apres-midi",16:"Apres-midi",17:"Apres-midi",
                    18:"Soir",19:"Soir",20:"Soir",21:"Soir",22:"Soir",23:"Soir"}
        age = body.age
        groupe_age = "Enfant" if age < 15 else "Adulte jeune" if age < 30 else "Adulte" if age < 60 else "Senior"

        new_row = {
            "ID_Urgence":          max_id_urg,
            "ID_Patient":          max_id_pat,
            "Nom_Complet":         body.nom_complet,
            "CIN":                 body.cin,
            "Age":                 body.age,
            "Sexe":                body.sexe,
            "Groupe_Sanguin":      body.groupe_sanguin,
            "Antecedents":         body.antecedents,
            "Etablissement":       body.etablissement,
            "Type_Etab":           "CHU",
            "Ville":               "Rabat",
            "Date_Arrivee":        dt.strftime("%Y-%m-%d %H:%M:%S"),
            "Date_Sortie":         date_sortie.strftime("%Y-%m-%d %H:%M:%S"),
            "Niveau_Triage":       body.niveau_triage,
            "Motif_Consultation":  body.motif_consultation,
            "Orientation":         body.orientation,
            "Duree_Sejour_min":    body.duree_sejour_min,
            "Nb_Medecins_Dispo":   body.nb_medecins_dispo,
            "Nb_Lits_Dispo":       body.nb_lits_dispo,
            "Jour_Ferie":          0,
            "Saison":              SAISONS[dt.month],
            "Heure_Arrivee":       dt.hour,
            "Jour_Semaine":        dt.weekday(),
            "Mois":                dt.month,
            "Annee":               dt.year,
            "Tranche_Horaire":     TRANCHES[dt.hour],
            "Nom_Jour":            JOURS_FR[dt.weekday()],
            "Groupe_Age":          groupe_age,
            "Est_Pic":             1 if dt.hour in range(8, 20) else 0,
            "Mutuelle":            body.mutuelle,
            "Prix_Sejour":         body.prix_sejour,
            "Prix_Soins":          body.prix_soins,
        }

        new_df = pd.DataFrame([new_row])

        if engine:
            # PostgreSQL : INSERT direct
            new_df["Date_Arrivee"] = pd.to_datetime(new_df["Date_Arrivee"])
            new_df["Date_Sortie"]  = pd.to_datetime(new_df["Date_Sortie"])
            new_df.to_sql("urgences", engine, if_exists="append", index=False)
        else:
            # Fallback CSV
            new_df.to_csv(urg_path, mode='a', header=False, index=False, encoding="utf-8-sig")

        # Recharger en mémoire
        reload_urg()

        return {"success": True, "id_urgence": new_row["ID_Urgence"], "id_patient": new_row["ID_Patient"]}
    except Exception as e:
        import traceback
        print("ADMISSION ERROR:", traceback.format_exc())
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# ALERTES CONFIGURABLES (seuils custom par utilisateur)
# ══════════════════════════════════════════════════════════════
_alert_config = {
    "duree_moy_seuil": 240,
    "taux_fugue_seuil": 3.0,
    "taux_p1_seuil": 5.0,
    "taux_hospit_seuil": 35.0,
}

class AlertConfig(BaseModel):
    duree_moy_seuil: float = 240
    taux_fugue_seuil: float = 3.0
    taux_p1_seuil: float = 5.0
    taux_hospit_seuil: float = 35.0

@app.get("/api/alertes/config")
def get_alert_config():
    return _alert_config

@app.post("/api/alertes/config")
def set_alert_config(cfg: AlertConfig):
    _alert_config.update(cfg.dict())
    return {"message": "Configuration mise à jour", "config": _alert_config}

@app.get("/api/alertes/check")
def check_alertes():
    try:
        urg = get_urg()
        total = len(urg)
        duree_moy   = round(float(urg["Duree_Sejour_min"].mean()), 1)
        taux_fugue  = round(len(urg[urg["Orientation"]=="Fugue"])/total*100, 2)
        taux_p1     = round(len(urg[urg["Niveau_Triage"].str.startswith("P1",na=False)])/total*100, 2)
        taux_hospit = round(len(urg[urg["Orientation"]=="Hospitalise"])/total*100, 2)
        alertes = []
        if duree_moy > _alert_config["duree_moy_seuil"]:
            alertes.append({"type": "Durée séjour", "valeur": f"{duree_moy} min",
                "seuil": f"{_alert_config['duree_moy_seuil']} min", "niveau": "critique"})
        if taux_fugue > _alert_config["taux_fugue_seuil"]:
            alertes.append({"type": "Taux de fugue", "valeur": f"{taux_fugue}%",
                "seuil": f"{_alert_config['taux_fugue_seuil']}%", "niveau": "warning"})
        if taux_p1 > _alert_config["taux_p1_seuil"]:
            alertes.append({"type": "Cas critiques P1", "valeur": f"{taux_p1}%",
                "seuil": f"{_alert_config['taux_p1_seuil']}%", "niveau": "critique"})
        if taux_hospit > _alert_config["taux_hospit_seuil"]:
            alertes.append({"type": "Taux hospitalisation", "valeur": f"{taux_hospit}%",
                "seuil": f"{_alert_config['taux_hospit_seuil']}%", "niveau": "warning"})
        return {"alertes": alertes, "stats": {"duree_moy": duree_moy, "taux_fugue": taux_fugue,
            "taux_p1": taux_p1, "taux_hospit": taux_hospit}, "config": _alert_config}
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# EMAIL RAPPORT
# ══════════════════════════════════════════════════════════════
class EmailInput(BaseModel):
    to: str
    subject: str = "Rapport CHU Ibn Sina — Dashboard Urgences"
    body: str = "Veuillez trouver ci-joint le rapport automatique du dashboard."
    pdf_base64: Optional[str] = None
    filename: str = "rapport_CHU.pdf"

@app.post("/api/email/send")
def send_email(payload: EmailInput):
    smtp_host = os.environ.get("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.environ.get("SMTP_PORT", 587))
    smtp_user = os.environ.get("SMTP_USER", "")
    smtp_pass = os.environ.get("SMTP_PASSWORD", "")

    if not smtp_user or not smtp_pass:
        raise HTTPException(400, detail="SMTP non configuré. Définissez SMTP_USER et SMTP_PASSWORD dans les variables d'environnement.")

    try:
        msg = MIMEMultipart()
        msg["From"]    = smtp_user
        msg["To"]      = payload.to
        msg["Subject"] = payload.subject
        msg.attach(MIMEText(payload.body, "plain", "utf-8"))

        if payload.pdf_base64:
            pdf_bytes = base64.b64decode(payload.pdf_base64)
            part = MIMEBase("application", "octet-stream")
            part.set_payload(pdf_bytes)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{payload.filename}"')
            msg.attach(part)

        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, payload.to, msg.as_string())

        return {"message": f"Email envoyé à {payload.to}"}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ══════════════════════════════════════════════════════════════
# PIPELINE ETL : Bronze → Gold
# ══════════════════════════════════════════════════════════════

@app.get("/api/pipeline/status")
def pipeline_status():
    """Retourne le nombre de lignes dans chaque couche."""
    if not engine:
        raise HTTPException(503, "PostgreSQL non disponible")
    try:
        with engine.connect() as conn:
            bronze_urg  = conn.execute(text("SELECT COUNT(*) FROM urgences_bronze")).scalar()
            bronze_soins= conn.execute(text("SELECT COUNT(*) FROM soins_bronze")).scalar()
            bronze_etab = conn.execute(text("SELECT COUNT(*) FROM etablissements_bronze")).scalar()
            gold_urg    = conn.execute(text('SELECT COUNT(*) FROM urgences')).scalar()
            gold_soins  = conn.execute(text('SELECT COUNT(*) FROM soins')).scalar()
            gold_etab   = conn.execute(text('SELECT COUNT(*) FROM etablissements')).scalar()
        return {
            "bronze": {"urgences": bronze_urg, "soins": bronze_soins, "etablissements": bronze_etab},
            "gold":   {"urgences": gold_urg,   "soins": gold_soins,   "etablissements": gold_etab},
        }
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/pipeline/run")
def run_pipeline(creds: HTTPAuthorizationCredentials = Depends(security)):
    """Transforme Bronze → Gold avec enrichissement des colonnes."""
    if not engine:
        raise HTTPException(503, "PostgreSQL non disponible")
    try:
        SAISONS   = {12:"Hiver",1:"Hiver",2:"Hiver",3:"Printemps",4:"Printemps",5:"Printemps",
                     6:"Ete",7:"Ete",8:"Ete",9:"Automne",10:"Automne",11:"Automne"}
        JOURS_FR  = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]
        TRANCHES  = {**{h:"Nuit"       for h in list(range(0,6))+[23]},
                     **{h:"Matin"      for h in range(6,12)},
                     **{h:"Apres-midi" for h in range(12,18)},
                     **{h:"Soir"       for h in range(18,23)}}

        # ── Charger Bronze ─────────────────────────────────────────
        df = pd.read_sql("SELECT * FROM urgences_bronze", engine)
        df["date_arrivee"] = pd.to_datetime(df["date_arrivee"], errors="coerce")
        df["date_sortie"]  = pd.to_datetime(df["date_sortie"],  errors="coerce")

        # ── Enrichissement → colonnes Gold ─────────────────────────
        df["Saison"]         = df["date_arrivee"].dt.month.map(SAISONS)
        df["Heure_Arrivee"]  = df["date_arrivee"].dt.hour
        df["Jour_Semaine"]   = df["date_arrivee"].dt.weekday
        df["Mois"]           = df["date_arrivee"].dt.month
        df["Annee"]          = df["date_arrivee"].dt.year
        df["Nom_Jour"]       = df["Jour_Semaine"].apply(lambda x: JOURS_FR[int(x)] if pd.notna(x) and 0 <= int(x) <= 6 else "")
        df["Tranche_Horaire"]= df["Heure_Arrivee"].apply(lambda h: TRANCHES.get(int(h), "Inconnu") if pd.notna(h) else "Inconnu")
        df["Groupe_Age"]     = df["age"].apply(
            lambda a: "Enfant" if pd.notna(a) and a < 15 else "Adulte jeune" if pd.notna(a) and a < 30 else "Adulte" if pd.notna(a) and a < 60 else "Senior" if pd.notna(a) else "Inconnu"
        )
        df["Est_Pic"]        = df["Heure_Arrivee"].apply(lambda h: 1 if pd.notna(h) and 8 <= h < 20 else 0)
        df["Type_Etab"]      = df["type_etab"]     if "type_etab"     in df.columns else "CHU"
        df["Ville"]          = df["ville"]         if "ville"         in df.columns else "Rabat"
        df["Jour_Ferie"]     = df["jour_ferie"]    if "jour_ferie"    in df.columns else 0
        df["Mutuelle"]       = df["mutuelle"]    if "mutuelle"    in df.columns else "Payant"
        df["Prix_Sejour"]    = df["prix_sejour"] if "prix_sejour" in df.columns else 0.0
        df["Prix_Soins"]     = df["prix_soins"]  if "prix_soins"  in df.columns else 0.0

        # ── Renommer colonnes bronze → gold ────────────────────────
        rename_map = {
            "id_urgence": "ID_Urgence", "id_patient": "ID_Patient",
            "nom_complet": "Nom_Complet", "age": "Age", "sexe": "Sexe",
            "cin": "CIN", "groupe_sanguin": "Groupe_Sanguin",
            "antecedents": "Antecedents", "etablissement": "Etablissement",
            "date_arrivee": "Date_Arrivee", "date_sortie": "Date_Sortie",
            "niveau_triage": "Niveau_Triage", "motif_consultation": "Motif_Consultation",
            "orientation": "Orientation", "duree_sejour_min": "Duree_Sejour_min",
            "nb_medecins_dispo": "Nb_Medecins_Dispo", "nb_lits_dispo": "Nb_Lits_Dispo",
        }
        df = df.rename(columns=rename_map)

        gold_cols = [
            "ID_Urgence","ID_Patient","Nom_Complet","Age","Sexe","CIN","Groupe_Sanguin",
            "Antecedents","Etablissement","Type_Etab","Ville","Date_Arrivee","Date_Sortie",
            "Niveau_Triage","Motif_Consultation","Orientation","Duree_Sejour_min",
            "Nb_Medecins_Dispo","Nb_Lits_Dispo","Jour_Ferie","Saison","Heure_Arrivee",
            "Jour_Semaine","Mois","Annee","Tranche_Horaire","Nom_Jour","Groupe_Age","Est_Pic",
            "Mutuelle","Prix_Sejour","Prix_Soins"
        ]
        df_gold = df[[c for c in gold_cols if c in df.columns]]

        # ── Insertion incrémentale : seulement les nouveaux enregistrements ──
        # (ne jamais écraser le Gold pour préserver les patients admis manuellement)
        with engine.connect() as _conn:
            existing_ids = set(
                pd.read_sql('SELECT DISTINCT "ID_Urgence" FROM urgences', engine)["ID_Urgence"].astype(str)
            )
        df_gold["_id_str"] = df_gold["ID_Urgence"].astype(str)
        df_new = df_gold[~df_gold["_id_str"].isin(existing_ids)].drop(columns=["_id_str"])
        df_gold.drop(columns=["_id_str"], inplace=True, errors="ignore")

        new_count = len(df_new)
        if new_count > 0:
            df_new.to_sql("urgences", engine, if_exists="append", index=False,
                          method="multi", chunksize=500)

        # Recharger le cache mémoire
        reload_urg()

        bronze_count = len(df)
        total_gold   = bronze_count  # après pipeline
        return {
            "success": True,
            "message": f"Pipeline terminé : {new_count} nouvelles lignes insérées ({bronze_count} Bronze, {total_gold} Gold total)",
            "bronze_rows": bronze_count,
            "new_rows":    new_count,
            "gold_rows":   total_gold,
        }
    except Exception as e:
        import traceback
        print("PIPELINE ERROR:", traceback.format_exc())
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# TEMPS RÉEL — KPIs DU JOUR
# ══════════════════════════════════════════════════════════════
@app.get("/api/kpis/live")
def get_kpis_live():
    """KPIs pour AUJOURD'HUI uniquement — temps réel."""
    try:
        urg = get_urg()
        today = pd.Timestamp.now().normalize()
        today_df = urg[urg["Date_Arrivee"].dt.normalize() == today]
        total_today = len(today_df)

        # Patients actifs (pas encore sortis) — NaN-safe
        now = pd.Timestamp.now()
        date_sortie = pd.to_datetime(urg["Date_Sortie"], errors="coerce")
        actifs = urg[
            (urg["Date_Arrivee"].dt.normalize() == today) &
            (date_sortie.isna() | (date_sortie >= now))
        ]

        taux_p1 = 0.0
        if total_today > 0:
            taux_p1 = round(
                len(today_df[today_df["Niveau_Triage"].str.startswith("P1", na=False)]) / total_today * 100, 1
            )

        # Charge globale : lits occupés / total lits (from etablissements)
        etab = data.get("etab", pd.DataFrame())
        total_lits = int(etab["capacite_lits"].sum()) if not etab.empty and "capacite_lits" in etab.columns else 500
        lits_occupes = min(len(actifs), total_lits)
        taux_charge = round(lits_occupes / total_lits * 100, 1) if total_lits > 0 else 0

        return {
            "patients_aujourd_hui": total_today,
            "patients_actifs":      len(actifs),
            "lits_occupes":         lits_occupes,
            "total_lits":           total_lits,
            "taux_charge":          taux_charge,
            "taux_p1_aujourd_hui":  taux_p1,
            "heure_maj":            now.strftime("%H:%M:%S"),
        }
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# TEMPS RÉEL — PATIENTS ACTUELS (AUJOURD'HUI)
# ══════════════════════════════════════════════════════════════
@app.get("/api/patients/aujourd_hui")
def get_patients_aujourd_hui(creds: HTTPAuthorizationCredentials = Depends(security)):
    """Retourne les patients admis aujourd'hui avec leur statut en temps réel."""
    try:
        urg = get_urg()
        today = pd.Timestamp.now().normalize()
        df = urg[urg["Date_Arrivee"].dt.normalize() == today].copy()
        # Filtre par établissement si directeur
        etab_user = _etab_from_creds(creds)
        if etab_user and "Etablissement" in df.columns:
            df = df[df["Etablissement"] == etab_user]
        df = df.sort_values("Date_Arrivee", ascending=False)

        # Charger les statuts depuis la table patient_statuts
        statuts: dict = {}
        if engine:
            try:
                with engine.connect() as conn:
                    rows = conn.execute(text("SELECT id_urgence, statut, lit_numero, updated_at, updated_by FROM patient_statuts")).fetchall()
                    for r in rows:
                        statuts[str(r[0])] = {"statut": r[1], "lit_numero": r[2] or "", "updated_at": r[3], "updated_by": r[4]}
            except Exception:
                pass

        results = []
        for _, row in df.iterrows():
            id_urg = str(row.get("ID_Urgence", ""))
            st_info = statuts.get(id_urg, {})
            results.append({
                "id_urgence":        id_urg,
                "nom_complet":       row.get("Nom_Complet", ""),
                "age":               row.get("Age", 0),
                "sexe":              row.get("Sexe", ""),
                "cin":               row.get("CIN", ""),
                "niveau_triage":     row.get("Niveau_Triage", ""),
                "motif":             row.get("Motif_Consultation", ""),
                "etablissement":     row.get("Etablissement", ""),
                "heure_arrivee":     row["Date_Arrivee"].strftime("%H:%M") if pd.notna(row["Date_Arrivee"]) else "",
                "statut":            st_info.get("statut", "En triage"),
                "lit_numero":        st_info.get("lit_numero", ""),
                "updated_at":        st_info.get("updated_at", ""),
                "updated_by":        st_info.get("updated_by", ""),
            })
        return results
    except Exception as e:
        raise HTTPException(500, str(e))


class StatutUpdate(BaseModel):
    statut:     str   # "En triage" | "En attente" | "En traitement" | "Sorti"
    lit_numero: str = ""   # ex: "A-12", "B-03"

@app.patch("/api/patients/{id_urgence}/statut")
def update_statut(id_urgence: str, body: StatutUpdate, creds: HTTPAuthorizationCredentials = Depends(security)):
    """Met à jour le statut et/ou le lit d'un patient en temps réel."""
    STATUTS_VALIDES = ["En triage", "En attente", "En traitement", "Sorti"]
    if body.statut not in STATUTS_VALIDES:
        raise HTTPException(400, f"Statut invalide. Valeurs acceptées : {STATUTS_VALIDES}")
    if not engine:
        raise HTTPException(503, "Base de données non disponible")
    try:
        payload = decode_jwt(creds.credentials) if creds else None
        username = payload.get("sub", "inconnu") if payload else "inconnu"
        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS patient_statuts (
                    id_urgence TEXT PRIMARY KEY,
                    statut     TEXT NOT NULL,
                    lit_numero TEXT DEFAULT '',
                    updated_at TEXT,
                    updated_by TEXT
                )
            """))
            conn.execute(text("""
                INSERT INTO patient_statuts (id_urgence, statut, lit_numero, updated_at, updated_by)
                VALUES (:id, :st, :lit, :at, :by)
                ON CONFLICT(id_urgence) DO UPDATE SET
                    statut=excluded.statut, lit_numero=excluded.lit_numero,
                    updated_at=excluded.updated_at, updated_by=excluded.updated_by
            """), {"id": id_urgence, "st": body.statut, "lit": body.lit_numero, "at": now, "by": username})
            conn.commit()
        return {"success": True, "id_urgence": id_urgence, "statut": body.statut, "lit_numero": body.lit_numero}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# PERSONNEL (Ressources Humaines)
# ══════════════════════════════════════════════════════════════
@app.get("/api/personnel")
def get_personnel(etablissement: str = "", role: str = "", statut: str = "",
                  creds: HTTPAuthorizationCredentials = Depends(security)):
    """Liste du personnel avec filtres optionnels."""
    etab_user = _etab_from_creds(creds)
    if etab_user and not etablissement:
        etablissement = etab_user
    if not engine:
        raise HTTPException(503, "DB non disponible")
    try:
        q = "SELECT * FROM personnel WHERE 1=1"
        params: dict = {}
        if etablissement:
            q += " AND etablissement = :etab"; params["etab"] = etablissement
        if role:
            q += " AND role = :role"; params["role"] = role
        if statut:
            q += " AND statut = :statut"; params["statut"] = statut
        q += " ORDER BY etablissement, role, nom_complet"
        with engine.connect() as conn:
            rows = conn.execute(text(q), params).fetchall()
            cols = conn.execute(text(q), params).keys() if False else \
                   ["id","matricule","etablissement","nom_complet","sexe","role",
                    "specialite","statut","telephone","email","date_embauche","updated_at"]
        return [dict(zip(cols, r)) for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/personnel/directeurs")
def get_directeurs():
    """Retourne un directeur (Chef de service) par établissement."""
    if not engine:
        raise HTTPException(503, "DB non disponible")
    try:
        with engine.connect() as conn:
            rows = conn.execute(text("""
                SELECT etablissement, matricule, nom_complet, sexe, role,
                       specialite, statut, telephone, email, date_embauche
                FROM personnel
                WHERE role = 'Chef de service'
                GROUP BY etablissement
                ORDER BY etablissement
            """)).fetchall()
        cols = ["etablissement","matricule","nom_complet","sexe","role",
                "specialite","statut","telephone","email","date_embauche"]
        return [dict(zip(cols, r)) for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/personnel/stats")
def get_personnel_stats(creds: HTTPAuthorizationCredentials = Depends(security)):
    """Statistiques RH par établissement."""
    if not engine:
        raise HTTPException(503, "DB non disponible")
    etab_user = _etab_from_creds(creds)
    try:
        with engine.connect() as conn:
            q = """
                SELECT etablissement,
                       COUNT(*) as total,
                       SUM(CASE WHEN role LIKE '%edecin%' OR role LIKE '%nterne%' THEN 1 ELSE 0 END) as medecins,
                       SUM(CASE WHEN role LIKE '%nfirmier%' OR role LIKE '%ide%' THEN 1 ELSE 0 END) as infirmiers,
                       SUM(CASE WHEN statut='En service' OR statut='En garde' THEN 1 ELSE 0 END) as en_service,
                       SUM(CASE WHEN statut='En conge' OR statut='Repos' THEN 1 ELSE 0 END) as absent
                FROM personnel
            """
            params: dict = {}
            if etab_user:
                q += " WHERE etablissement = :etab"; params["etab"] = etab_user
            q += " GROUP BY etablissement ORDER BY total DESC"
            rows = conn.execute(text(q), params).fetchall()
        cols = ["etablissement","total","medecins","infirmiers","en_service","absent"]
        return [dict(zip(cols, r)) for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))

class PersonnelUpdate(BaseModel):
    statut: str

@app.patch("/api/personnel/{matricule}")
def update_personnel_statut(matricule: str, body: PersonnelUpdate, creds: HTTPAuthorizationCredentials = Depends(security)):
    """Met à jour le statut d'un membre du personnel."""
    STATUTS_OK = ["En service","En garde","En conge","Repos"]
    if body.statut not in STATUTS_OK:
        raise HTTPException(400, f"Statut invalide : {STATUTS_OK}")
    if not engine:
        raise HTTPException(503, "DB non disponible")
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    with engine.connect() as conn:
        conn.execute(text("UPDATE personnel SET statut=:s, updated_at=:u WHERE matricule=:m"),
                     {"s": body.statut, "u": now, "m": matricule})
        conn.commit()
    return {"success": True}


# ══════════════════════════════════════════════════════════════
# LITS (table complète)
# ══════════════════════════════════════════════════════════════
@app.get("/api/lits")
def get_lits(etablissement: str = "", service: str = "", statut: str = "",
             creds: HTTPAuthorizationCredentials = Depends(security)):
    """Liste des lits avec filtres optionnels."""
    etab_user = _etab_from_creds(creds)
    if etab_user and not etablissement:
        etablissement = etab_user
    if not engine:
        raise HTTPException(503, "DB non disponible")
    try:
        q = "SELECT * FROM lits WHERE 1=1"
        params: dict = {}
        if etablissement:
            q += " AND etablissement = :etab"; params["etab"] = etablissement
        if service:
            q += " AND service = :service"; params["service"] = service
        if statut:
            q += " AND statut = :statut"; params["statut"] = statut
        q += " ORDER BY etablissement, service, numero_lit"
        with engine.connect() as conn:
            rows = conn.execute(text(q), params).fetchall()
        cols = ["id","etablissement","numero_lit","service","type_lit","statut",
                "id_patient","nom_patient","date_admission","updated_at"]
        return [dict(zip(cols, r)) for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/lits/stats")
def get_lits_stats(creds: HTTPAuthorizationCredentials = Depends(security)):
    """Statistiques des lits par établissement."""
    if not engine:
        raise HTTPException(503, "DB non disponible")
    etab_user = _etab_from_creds(creds)
    try:
        with engine.connect() as conn:
            q = """
                SELECT etablissement,
                       COUNT(*) as total,
                       SUM(CASE WHEN statut='Disponible' THEN 1 ELSE 0 END) as disponibles,
                       SUM(CASE WHEN statut='Occupe' THEN 1 ELSE 0 END) as occupes,
                       SUM(CASE WHEN statut='En maintenance' THEN 1 ELSE 0 END) as maintenance,
                       ROUND(SUM(CASE WHEN statut='Occupe' THEN 1.0 ELSE 0 END)*100/COUNT(*),1) as taux_occupation
                FROM lits
            """
            params: dict = {}
            if etab_user:
                q += " WHERE etablissement = :etab"; params["etab"] = etab_user
            q += " GROUP BY etablissement ORDER BY taux_occupation DESC"
            rows = conn.execute(text(q), params).fetchall()
        cols = ["etablissement","total","disponibles","occupes","maintenance","taux_occupation"]
        return [dict(zip(cols, r)) for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/anomalies")
def get_anomalies(creds: HTTPAuthorizationCredentials = Depends(security)):
    """Compare le flux horaire du dernier jour disponible vs la moyenne historique."""
    try:
        urg = get_urg()
        etab_user = _etab_from_creds(creds)
        if etab_user:
            urg = urg[urg["Etablissement"] == etab_user]
        urg = urg.copy()
        urg["heure"] = urg["Date_Arrivee"].dt.hour
        urg["date"]  = urg["Date_Arrivee"].dt.normalize()
        n_days = max(urg["date"].nunique(), 1)
        historical = urg.groupby("heure").size() / n_days
        last_date = urg["date"].max()
        last_df   = urg[urg["date"] == last_date]
        last_hourly = last_df.groupby("heure").size()
        result = []
        for h in range(24):
            hist_avg  = round(float(historical.get(h, 0)), 1)
            today_val = int(last_hourly.get(h, 0))
            ecart_pct = round((today_val - hist_avg) / (hist_avg + 0.01) * 100, 1)
            result.append({
                "heure":          h,
                "historique_moy": hist_avg,
                "aujourd_hui":    today_val,
                "ecart_pct":      ecart_pct,
                "anomalie":       abs(ecart_pct) > 30,
            })
        return result
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/lits/recommander")
def recommander_lits(service: str = "", creds: HTTPAuthorizationCredentials = Depends(security)):
    """Retourne les établissements avec lits disponibles, triés par disponibilité."""
    if not engine:
        raise HTTPException(503, "DB non disponible")
    try:
        with engine.connect() as conn:
            q = """
                SELECT etablissement, service,
                       COUNT(*) as total,
                       SUM(CASE WHEN statut='Disponible' THEN 1 ELSE 0 END) as disponibles,
                       ROUND(SUM(CASE WHEN statut='Occupe' THEN 1.0 ELSE 0 END)*100/COUNT(*), 1) as taux_occupation
                FROM lits WHERE 1=1
            """
            params: dict = {}
            if service:
                q += " AND LOWER(service) LIKE :svc"; params["svc"] = f"%{service.lower()}%"
            q += " GROUP BY etablissement, service HAVING disponibles > 0 ORDER BY disponibles DESC LIMIT 20"
            rows = conn.execute(text(q), params).fetchall()
        cols = ["etablissement","service","total","disponibles","taux_occupation"]
        return [dict(zip(cols, r)) for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))


class LitUpdate(BaseModel):
    statut:         str
    id_patient:     str = ""
    nom_patient:    str = ""

@app.patch("/api/lits/{etablissement}/{numero_lit}")
def update_lit(etablissement: str, numero_lit: str, body: LitUpdate,
               creds: HTTPAuthorizationCredentials = Depends(security)):
    """Met à jour le statut d'un lit (Disponible/Occupé/En maintenance)."""
    STATUTS_OK = ["Disponible","Occupe","En maintenance","Reserve"]
    if body.statut not in STATUTS_OK:
        raise HTTPException(400, f"Statut invalide : {STATUTS_OK}")
    if not engine:
        raise HTTPException(503, "DB non disponible")
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    date_adm = now if body.statut == "Occupe" else ""
    with engine.connect() as conn:
        conn.execute(text("""
            UPDATE lits SET statut=:s, id_patient=:ip, nom_patient=:np,
                           date_admission=:da, updated_at=:u
            WHERE etablissement=:e AND numero_lit=:n
        """), {"s": body.statut, "ip": body.id_patient, "np": body.nom_patient,
               "da": date_adm, "u": now, "e": etablissement, "n": numero_lit})
        conn.commit()
    return {"success": True}


# ══════════════════════════════════════════════════════════════
# LITS DISPONIBLES (endpoint simplifié pour dropdown)
# ══════════════════════════════════════════════════════════════
@app.get("/api/lits/disponibles")
def get_lits_disponibles(etablissement: str = ""):
    """Retourne la liste des lits disponibles depuis la table lits."""
    if not engine:
        raise HTTPException(503, "DB non disponible")
    try:
        params: dict = {}
        if etablissement:
            q = "SELECT numero_lit FROM lits WHERE statut='Disponible' AND etablissement=:e ORDER BY service, numero_lit"
            params["e"] = etablissement
        else:
            q = "SELECT numero_lit FROM lits WHERE statut='Disponible' ORDER BY etablissement, service, numero_lit"
        with engine.connect() as conn:
            disponibles = [r[0] for r in conn.execute(text(q), params).fetchall()]
            q2 = "SELECT numero_lit FROM lits WHERE statut='Occupe'"
            if etablissement:
                q2 += " AND etablissement=:e"
            occupes = [r[0] for r in conn.execute(text(q2), params).fetchall()]
            q3 = "SELECT COUNT(*) FROM lits" + (" WHERE etablissement=:e" if etablissement else "")
            total = conn.execute(text(q3), params).scalar()
        return {
            "disponibles": disponibles,
            "occupes":     occupes,
            "total":       total,
            "nb_dispo":    len(disponibles),
        }
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════
# GESTION DES UTILISATEURS
# ══════════════════════════════════════════════════════════════
def _init_users_table():
    """Crée la table users et insère les comptes par défaut si vide."""
    if not engine:
        return
    with engine.connect() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS users (
                username       TEXT PRIMARY KEY,
                password       TEXT NOT NULL,
                role           TEXT NOT NULL,
                nom_complet    TEXT DEFAULT '',
                actif          INTEGER DEFAULT 1,
                created_at     TEXT,
                etablissement  TEXT DEFAULT '',
                password_plain TEXT DEFAULT ''
            )
        """))
        # Migrations pour colonnes ajoutées
        for col, definition in [("etablissement", "TEXT DEFAULT ''"), ("password_plain", "TEXT DEFAULT ''")]:
            try:
                conn.execute(text(f"ALTER TABLE users ADD COLUMN {col} {definition}"))
            except Exception:
                pass

        count = conn.execute(text("SELECT COUNT(*) FROM users")).scalar()
        if count == 0:
            now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            defaults = [
                ("admin",        _hash("admin123"), "admin",        "Administrateur",        "admin123"),
                ("chef_medecin", _hash("chef123"),  "chef_medecin", "Médecin Chef",           "chef123"),
                ("urgentiste",   _hash("urg123"),   "urgentiste",   "Médecin Urgentiste",     "urg123"),
                ("infirmier",    _hash("inf123"),   "infirmier",    "Cadre Infirmier",        "inf123"),
                ("directeur",    _hash("dir123"),   "directeur",    "Directeur Médical",      "dir123"),
                ("analyste",     _hash("ana123"),   "analyste",     "Data Analyst",           "ana123"),
                ("admin_si",     _hash("si123"),    "admin_si",     "Admin SI",               "si123"),
            ]
            for u, p, r, n, pp in defaults:
                conn.execute(text(
                    "INSERT OR IGNORE INTO users (username, password, role, nom_complet, actif, created_at, etablissement, password_plain) VALUES (:u,:p,:r,:n,1,:at,'', :pp)"
                ), {"u": u, "p": p, "r": r, "n": n, "at": now, "pp": pp})
        conn.commit()


def _get_user_from_db(username: str):
    """Récupère un utilisateur depuis la DB (fallback sur USERS dict si DB KO)."""
    if engine:
        try:
            with engine.connect() as conn:
                row = conn.execute(
                    text("SELECT username, password, role, actif, etablissement FROM users WHERE username=:u"),
                    {"u": username}
                ).fetchone()
                if row:
                    return {"username": row[0], "password": row[1], "role": row[2], "actif": row[3], "etablissement": row[4] or ""}
        except Exception:
            pass
    u = USERS.get(username)
    if u:
        return {**u, "etablissement": ""}
    return None


# Remplacer la logique de login pour utiliser la DB
@app.post("/api/auth/login")
def login_v2(body: LoginInput):
    hashed = _hash(body.password)
    u = _get_user_from_db(body.username)
    if not u or u["password"] != hashed:
        raise HTTPException(status_code=401, detail="Identifiants incorrects")
    if isinstance(u.get("actif"), int) and u["actif"] == 0:
        raise HTTPException(status_code=403, detail="Compte désactivé")
    etab = u.get("etablissement", "")
    token = create_jwt(body.username, u["role"], etab)
    return {"token": token, "username": body.username, "role": u["role"], "etablissement": etab}


class UserCreate(BaseModel):
    username:   str
    password:   str
    role:       str
    nom_complet: str = ""

class UserUpdate(BaseModel):
    password:   str = ""
    role:       str = ""
    nom_complet: str = ""
    actif:      int = 1

ROLES_VALIDES = ["admin","chef_medecin","urgentiste","infirmier","directeur","analyste","admin_si"]

def _require_admin(creds: HTTPAuthorizationCredentials):
    payload = decode_jwt(creds.credentials) if creds else None
    if not payload or payload.get("role") != "admin":
        raise HTTPException(403, "Accès réservé à l'administrateur")
    return payload

def _etab_from_creds(creds: HTTPAuthorizationCredentials) -> str:
    """Retourne l'établissement du user connecté ('' = pas de filtre)."""
    payload = decode_jwt(creds.credentials) if creds else None
    if not payload:
        return ""
    return payload.get("etablissement", "") or ""

DEFAULT_DIR_PASSWORD = "DirecteurCHU2025"

@app.post("/api/users/seed-directors")
def seed_director_accounts(creds: HTTPAuthorizationCredentials = Depends(security)):
    """Crée un compte de connexion pour chaque Chef de service de la table personnel."""
    _require_admin(creds)
    if not engine:
        raise HTTPException(503, "DB non disponible")
    try:
        with engine.connect() as conn:
            chefs = conn.execute(text("""
                SELECT matricule, nom_complet, etablissement, telephone, email
                FROM personnel
                WHERE role = 'Chef de service'
                GROUP BY etablissement
                ORDER BY etablissement
            """)).fetchall()

            now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            created, updated, skipped = [], [], []

            for chef in chefs:
                matricule, nom_complet, etablissement, telephone, email = chef
                nom_complet = (nom_complet or "").strip()

                # Username: initiale_prenom.nom en minuscules, sans accents
                import unicodedata
                def normalize(s):
                    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

                parts = nom_complet.split()
                if len(parts) >= 2:
                    prenom_init = parts[0][0].lower() if parts[0] not in ('Dr.', 'Pr.', 'M.', 'Mme') else (parts[1][0].lower() if len(parts) > 1 else 'x')
                    nom_famille = parts[-1].lower()
                    base_username = normalize(prenom_init + '.' + nom_famille)
                else:
                    base_username = normalize(nom_complet.lower().replace(' ', '_'))

                # Mot de passe par défaut simple et connu
                hashed_pwd = _hash(DEFAULT_DIR_PASSWORD)

                # Vérifier si un compte directeur existe déjà pour cet établissement
                existing_etab = conn.execute(
                    text("SELECT username FROM users WHERE role='directeur' AND etablissement=:e"), {"e": etablissement}
                ).fetchone()

                if existing_etab:
                    # Mettre à jour le compte existant pour cet établissement
                    username = existing_etab[0]
                    conn.execute(text(
                        "UPDATE users SET password=:p, password_plain=:pp, nom_complet=:n, actif=1 WHERE username=:u"
                    ), {"p": hashed_pwd, "pp": DEFAULT_DIR_PASSWORD, "n": nom_complet, "u": username})
                    updated.append({
                        "username":     username,
                        "mot_de_passe": DEFAULT_DIR_PASSWORD,
                        "nom_complet":  nom_complet,
                        "etablissement": etablissement,
                    })
                else:
                    # Générer un username unique si collision
                    username = base_username
                    suffix = 2
                    while conn.execute(text("SELECT 1 FROM users WHERE username=:u"), {"u": username}).fetchone():
                        username = f"{base_username}{suffix}"
                        suffix += 1

                    conn.execute(text("""
                        INSERT INTO users (username, password, password_plain, role, nom_complet, actif, created_at, etablissement)
                        VALUES (:u, :p, :pp, 'directeur', :n, 1, :at, :e)
                    """), {"u": username, "p": hashed_pwd, "pp": DEFAULT_DIR_PASSWORD, "n": nom_complet, "at": now, "e": etablissement})
                    created.append({
                        "username":     username,
                        "mot_de_passe": DEFAULT_DIR_PASSWORD,
                        "nom_complet":  nom_complet,
                        "etablissement": etablissement,
                    })

            conn.commit()

        return {
            "created":       created,
            "updated":       updated,
            "skipped":       skipped,
            "total_created": len(created),
            "total_updated": len(updated),
            "mot_de_passe_defaut": DEFAULT_DIR_PASSWORD,
        }
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/users")
def list_users(creds: HTTPAuthorizationCredentials = Depends(security)):
    _require_admin(creds)
    if not engine:
        raise HTTPException(503, "DB non disponible")
    with engine.connect() as conn:
        rows = conn.execute(text(
            "SELECT username, role, nom_complet, actif, created_at, etablissement, password_plain FROM users ORDER BY role, username"
        )).fetchall()
    return [{"username": r[0], "role": r[1], "nom_complet": r[2], "actif": r[3],
             "created_at": r[4], "etablissement": r[5] or "", "password_plain": r[6] or ""} for r in rows]

class UserCreate(BaseModel):
    username:     str
    password:     str
    role:         str
    nom_complet:  str = ""
    etablissement: str = ""

@app.post("/api/users")
def create_user(body: UserCreate, creds: HTTPAuthorizationCredentials = Depends(security)):
    _require_admin(creds)
    if body.role not in ROLES_VALIDES:
        raise HTTPException(400, f"Rôle invalide : {body.role}")
    if not engine:
        raise HTTPException(503, "DB non disponible")
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with engine.connect() as conn:
            existing = conn.execute(text("SELECT username FROM users WHERE username=:u"), {"u": body.username}).fetchone()
            if existing:
                raise HTTPException(409, "Utilisateur déjà existant")
            conn.execute(text(
                "INSERT INTO users (username, password, role, nom_complet, actif, created_at, etablissement, password_plain) VALUES (:u,:p,:r,:n,1,:at,:e,:pp)"
            ), {"u": body.username, "p": _hash(body.password), "r": body.role, "n": body.nom_complet,
                "at": now, "e": body.etablissement, "pp": body.password})
            conn.commit()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))
    return {"success": True, "username": body.username}

class UserUpdate(BaseModel):
    password:      str = ""
    role:          str = ""
    nom_complet:   str = ""
    actif:         int = 1
    etablissement: str = ""

@app.patch("/api/users/{username}")
def update_user(username: str, body: UserUpdate, creds: HTTPAuthorizationCredentials = Depends(security)):
    _require_admin(creds)
    if not engine:
        raise HTTPException(503, "DB non disponible")
    with engine.connect() as conn:
        if body.password:
            conn.execute(text("UPDATE users SET password=:p, password_plain=:pp WHERE username=:u"),
                         {"p": _hash(body.password), "pp": body.password, "u": username})
        if body.role and body.role in ROLES_VALIDES:
            conn.execute(text("UPDATE users SET role=:r WHERE username=:u"), {"r": body.role, "u": username})
        if body.nom_complet:
            conn.execute(text("UPDATE users SET nom_complet=:n WHERE username=:u"), {"n": body.nom_complet, "u": username})
        conn.execute(text("UPDATE users SET actif=:a, etablissement=:e WHERE username=:u"),
                     {"a": body.actif, "e": body.etablissement, "u": username})
        conn.commit()
    return {"success": True}

@app.delete("/api/users/{username}")
def delete_user(username: str, creds: HTTPAuthorizationCredentials = Depends(security)):
    _require_admin(creds)
    if username == "admin":
        raise HTTPException(400, "Impossible de supprimer le compte admin principal")
    if not engine:
        raise HTTPException(503, "DB non disponible")
    with engine.connect() as conn:
        conn.execute(text("DELETE FROM users WHERE username=:u"), {"u": username})
        conn.commit()
    return {"success": True}
